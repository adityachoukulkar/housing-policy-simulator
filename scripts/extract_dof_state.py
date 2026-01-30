#!/usr/bin/env python3
"""Extract statewide population and housing units from DOF E-5/E-8 Excel files."""

from __future__ import annotations

import csv
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, Tuple

from openpyxl import load_workbook


def extract_e5_state(path: Path) -> Dict[int, Tuple[float, float]]:
    out: Dict[int, Tuple[float, float]] = {}
    wb = load_workbook(path, data_only=True)
    for sheet in wb.worksheets:
        if not sheet.title.startswith("E5CountyState"):
            continue
        year = int(sheet.title.replace("E5CountyState", ""))
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) < 5:
            continue
        header = rows[3]
        # Expect: County, Total(pop), Household, Group Quarters, Total(housing), ...
        try:
            pop_total_idx = header.index("Total")
            house_total_idx = header.index("Total", pop_total_idx + 1)
        except ValueError:
            # fallback to fixed indices
            pop_total_idx = 1
            house_total_idx = 4
        for row in rows[4:]:
            if not row or str(row[0]).strip().lower() != "california":
                continue
            pop = row[pop_total_idx]
            house = row[house_total_idx]
            if pop is None or house is None:
                continue
            out[year] = (float(pop), float(house))
            break
    return out


def extract_e8_state(path: Path) -> Dict[int, Tuple[float, float]]:
    out: Dict[int, Tuple[float, float]] = {}
    wb = load_workbook(path, data_only=True)
    if "E-8 by geography" not in wb.sheetnames:
        return out
    sheet = wb["E-8 by geography"]
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 5:
        return out

    header = rows[2]
    # columns: County, City, Date, Total(pop), Household, Group Quarters, Total(housing), ...
    try:
        date_idx = header.index("Date")
        pop_total_idx = header.index("Total")
        house_total_idx = header.index("Total", pop_total_idx + 1)
        county_idx = header.index("County")
        city_idx = header.index("City")
    except ValueError:
        return out

    for row in rows[3:]:
        if not row or len(row) <= max(date_idx, house_total_idx, city_idx):
            continue
        county = str(row[county_idx]).strip().lower()
        city = str(row[city_idx]).strip().lower()
        if county != "california" or city != "state total":
            continue
        date_val = row[date_idx]
        if not isinstance(date_val, datetime):
            continue
        year = date_val.year
        pop = row[pop_total_idx]
        house = row[house_total_idx]
        if pop is None or house is None:
            continue
        out[year] = (float(pop), float(house))
    return out


def write_output(path: Path, series: Dict[int, Tuple[float, float]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    years = sorted(series.keys())
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["year", "population", "housing_units"])
        writer.writeheader()
        for y in years:
            pop, house = series[y]
            writer.writerow({"year": y, "population": pop, "housing_units": house})


def main() -> int:
    e5_path = Path("data/raw/dof_e5/e5_state.xlsx")
    e8_path = Path("data/raw/dof_e5/e8_state.xlsx")

    if not e5_path.exists() and not e8_path.exists():
        print("Missing DOF Excel files. Run download_data.py first.", file=sys.stderr)
        return 1

    series: Dict[int, Tuple[float, float]] = {}
    if e8_path.exists():
        series.update(extract_e8_state(e8_path))
    if e5_path.exists():
        series.update(extract_e5_state(e5_path))

    if not series:
        print("Failed to extract DOF state series.", file=sys.stderr)
        return 1

    write_output(Path("data/raw/dof_e5/e5_state.csv"), series)
    print("Wrote data/raw/dof_e5/e5_state.csv")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
